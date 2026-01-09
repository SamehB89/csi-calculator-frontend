import pandas as pd
import openpyxl

# Load Excel file to examine structure
print("="*100)
print("EXAMINING CSI.XLSM STRUCTURE IN DETAIL")
print("="*100)

# Load with openpyxl to see raw structure
wb = openpyxl.load_workbook('CSI.xlsm')
ws = wb.active

print(f"\nWorksheet name: {ws.title}")
print(f"Total rows: {ws.max_row}")
print(f"Total columns: {ws.max_column}")

print("\n" + "="*100)
print("FIRST 20 ROWS (RAW DATA)")
print("="*100)

for i in range(1, min(21, ws.max_row + 1)):
    row_data = []
    for j in range(1, min(11, ws.max_column + 1)):  # First 10 columns
        cell = ws.cell(row=i, column=j)
        row_data.append(str(cell.value)[:30] if cell.value else "")
    print(f"Row {i:3d}: {' | '.join(row_data)}")

# Now load with pandas to see structure
df = pd.read_excel('CSI.xlsm', engine='openpyxl')

print("\n" + "="*100)
print("PANDAS INTERPRETATION (First 10 rows)")
print("="*100)
print(df.head(10).to_string())

# Look for WOOD & PLASTIC specifically
print("\n" + "="*100)
print("SEARCHING FOR 'WOOD' or 'DOOR' RELATED ENTRIES")
print("="*100)

for idx, row in df.iterrows():
    row_str = ' | '.join([str(x)[:50] for x in row.values[:10]])
    if 'WOOD' in row_str.upper() or 'DOOR' in row_str.upper():
        print(f"\nRow {idx}: {row_str}")
        if idx > 50 and idx < 200:  # Limit search range
            # Show some context
            print(f"  Previous row {idx-1}: {' | '.join([str(x)[:30] for x in df.iloc[idx-1].values[:5]])}")
            print(f"  Next row {idx+1}: {' | '.join([str(x)[:30] for x in df.iloc[idx+1].values[:5]])}")

# Check how divisions are organized
print("\n" + "="*100)
print("ANALYZING DIVISION STRUCTURE")
print("="*100)

# Read with different header settings
df_no_header = pd.read_excel('CSI.xlsm', engine='openpyxl', header=None)
print("\nFirst 30 rows without header interpretation:")
for i in range(30):
    print(f"Row {i}: {df_no_header.iloc[i, 0:8].values}")

wb.close()
